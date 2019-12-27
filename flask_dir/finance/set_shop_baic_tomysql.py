import openpyxl
import pymysql
import uuid

"""
把财务的门店基础信息导入到数据库的门店基础信息表里面
"""
filepath='C:\\Users\\tl\\Documents\\门店基础信息表for技术-20191128.xlsx'
wb = openpyxl.load_workbook(filepath)

sheetnames = wb.sheetnames

#累计营业利润

for sheet_name in sheetnames:

    ws=wb[sheet_name]

    max_column=ws.max_column
    all_data = []
    for column_index in range(2,101 ):
        list=[]
        shop_name = ws.cell(row=column_index, column=1).value
        shop_id = uuid.uuid3(uuid.NAMESPACE_DNS, str(shop_name))
        open_data = ws.cell(row=column_index, column=2).value
        type1 = ws.cell(row=column_index, column=3).value
        region = ws.cell(row=column_index, column=4).value
        city = ws.cell(row=column_index, column=5).value
        source = ws.cell(row=column_index, column=6).value
        state = ws.cell(row=column_index, column=7).value
        list.append(shop_id)
        list.append(shop_name)
        list.append(open_data)
        list.append(type1)
        list.append(region)
        list.append(city)
        list.append(source)
        list.append(state)
        all_data.append(tuple(list))

    sql_select = "INSERT IGNORE into `门店基础信息`  (`门店id`,`门店名称`,`开业日期`,`类型`,`区域`,`城市`,`项目来源`,`项目状态`) VALUES "

    db = pymysql.connect(host="bj-cdb-cwu7v42u.sql.tencentcdb.com", user="root",
                         password="xmxc1234", db="manage_info", port=62864)
    cur = db.cursor()

    for data in all_data:
        id = data[0]
        shop_name = data[2]

        shop_date = data[1]
        print(id)
        sql_select += " ('%s','%s','%s','%s','%s','%s','%s','%s')," % data


    sql_select = sql_select.rstrip(',')
    cur.execute(sql_select)
    db.commit()  # 事务提交
    cur.close()
    db.close()

    # try:
    #
    #
    # except Exception as e:
    #     db.rollback()  # 事务回滚
    # else:
    #     db.commit()  # 事务提交
    #     cur.close()
    #     db.close()