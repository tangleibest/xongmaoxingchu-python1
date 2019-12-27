# -*- coding: gbk -*-
import json

from flask import Flask, request,render_template,Response
import os

import pymysql
import uuid
import xlwt
import time
from openpyxl import load_workbook

"""
���񹤾߽ӿ�
"""

def set_style(font_name,font_colour,font_size,font_bold,underline,pattern_colour,ali_horz,ali_vert,left,right,top,bottom,italic):
    # ���ñ�����ɫ
    pattern = xlwt.Pattern()
    # ���ñ�����ɫ��ģʽ
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    # ������ɫ
    pattern.pattern_fore_colour = int(pattern_colour)
    font = xlwt.Font()
    # ��������
    font.name = font_name
    # ������ɫ
    font.colour_index = int(font_colour)
    # �����С��11Ϊ�ֺţ�20Ϊ������λ
    font.height = 20 * int(font_size)
    font.bold = font_bold
    #б��
    font.italic = italic
    # �»���
    font.underline = underline
    # ���õ�Ԫ����뷽ʽ
    alignment = xlwt.Alignment()
    # 0x01(��˶���)��0x02(ˮƽ�����Ͼ��ж���)��0x03(�Ҷ˶���)
    alignment.horz = ali_horz
    # 0x00(�϶˶���)�� 0x01(��ֱ�����Ͼ��ж���)��0x02(�׶˶���)
    alignment.vert = ali_vert

    borders = xlwt.Borders()
    # ϸʵ��:1��С��ʵ��:2��ϸ����:3����ϸ����:4�����ʵ��:5��˫��:6��ϸ������:7
    # �������:8��ϸ�㻮��:9���ֵ㻮��:10��ϸ˫�㻮��:11����˫�㻮��:12��б�㻮��:13
    borders.left = left
    borders.right = right
    borders.top = top
    borders.bottom = bottom

    # ��ʼ����ʽ
    style0 = xlwt.XFStyle()
    style0.font = font
    style0.pattern = pattern
    style0.alignment = alignment
    style0.borders = borders
    return style0
def file_iterator(file_path, chunk_size=512):
    """
        �ļ���ȡ������
    :param file_path:�ļ�·��
    :param chunk_size: ÿ�ζ�ȡ����С
    :return:
    """
    with open(file_path, 'rb') as target_file:
        while True:
            chunk = target_file.read(chunk_size)
            if chunk:
                yield chunk
            else:
                break
def to_json(obj):
    """
        ����
    :return:
    """
    return json.dumps(obj, ensure_ascii=False)
app = Flask(__name__)
app.secret_key = 'some_secret'

#�����ϴ��ŵ꿨Ƭҳ��
@app.route('/upload_card_data')
def upload_card_data():
    return render_template("commit.html")
#�����ϴ�Դ����ҳ��
@app.route('/source_data')
def source_data():
    return render_template("commit2.html")
#���ؽ����Ƭ��ͼ
@app.route('/download_card')
def download_card():
    str_request = request.args.get('shop')
    if str_request is None:
        return to_json({'success': 0, 'message': '���������'})
    else:
        if '��' in str_request :
            return to_json({'success': 0, 'message': "�������зǷ�������,���'��'��Ϊ','"})
        else:

                book = xlwt.Workbook(encoding='utf-8')
                db = pymysql.connect(host="bj-cdb-cwu7v42u.sql.tencentcdb.com", user="root",
                                     password="xmxc1234", db="test", port=62864)
                cur = db.cursor()
                cur1 = db.cursor()
                cur2 = db.cursor()

                splited = str(str_request).split(',')
                where_str = ''
                for every_splited in splited:
                    where_str += "'%s'," % every_splited
                where_str = where_str.strip(',')
                sql_select = "SELECT  * from `�ŵ���Ϣ` where `�ŵ�����` in (%s)" % where_str
                # sql_select = "SELECT  * from `�ŵ���Ϣ` order by `�ŵ�����` "

                cur.execute(sql_select)

                results = cur.fetchall()

                """
                �ظ���Ԫ����ʽ 
                """
                style1 = set_style('����', 0, 8, False, False, 1, 0x01, 0x01, 1, 1, 0, 0, False)
                style2 = set_style('����', 0, 8, False, False, 1, 0x03, 0x01, 1, 1, 0, 0, False)
                style3 = set_style('����', 0, 8, False, False, 9, 0x01, 0x01, 0, 0, 0, 0, False)
                style4 = set_style('����', 0, 8, False, False, 9, 0x02, 0x01, 0, 0, 0, 0, False)
                style5 = set_style('����', 0, 8, True, False, 43, 0x01, 0x01, 0, 0, 0, 0, False)
                style6 = set_style('����', 0, 8, False, False, 43, 0x01, 0x01, 0, 0, 0, 0, False)
                style7 = set_style('����', 0, 8, True, True, 1, 0x01, 0x01, 0, 0, 0, 0, False)
                style8 = set_style('����', 10, 8, True, True, 1, 0x02, 0x01, 1, 1, 1, 1, False)
                style9 = set_style('����', 0, 8, False, False, 1, 0x01, 0x01, 1, 1, 0, 1, False)
                style10 = set_style('����', 1, 8, True, False, 10, 0x01, 0x01, 0, 0, 0, 0, False)
                style11 = set_style('����', 1, 16, False, True, 10, 0x01, 0x01, 0, 0, 0, 0, False)
                style12 = set_style('����', 1, 16, False, False, 10, 0x01, 0x01, 0, 0, 0, 0, False)
                style13 = set_style('����', 2, 8, True, False, 1, 0x01, 0x01, 1, 1, 0, 1, False)
                style14 = set_style('����', 2, 8, True, False, 1, 0x03, 0x01, 1, 1, 0, 1, False)
                style15 = set_style('����', 2, 8, True, False, 1, 0x01, 0x01, 1, 1, 1, 1, False)
                style16 = set_style('����', 2, 8, True, False, 1, 0x03, 0x01, 1, 1, 1, 1, False)
                style17 = set_style('����', 1, 8, True, False, 10, 0x02, 0x01, 0, 0, 0, 0, False)
                style18 = set_style('����', 0, 8, False, True, 43, 0x02, 0x01, 0, 0, 0, 0, True)
                style19 = set_style('����', 10, 8, False, False, 1, 0x02, 0x01, 0, 0, 0, 0, True)
                style20 = set_style('����', 0, 8, False, False, 43, 0x02, 0x01, 0, 0, 0, 0, False)
                style21 = set_style('����', 10, 8, True, False, 43, 0x02, 0x01, 0, 0, 0, 0, True)
                style22 = set_style('����', 0, 8, False, False, 24, 0x01, 0x01, 0, 0, 0, 0, False)

                for row in results:

                    shop_all_cost2 = row[29]

                    shop_name = row[1]
                    sql_select1 = "SELECT  * from `��������¼���` where `�ŵ�id`='%s'" % row[0]

                    cur1.execute(sql_select1)
                    results1 = cur1.fetchall()

                    i = 0

                    sheet = book.add_sheet(shop_name, cell_overwrite_ok=True)
                    # ������ֱ���Exception: Attempt to overwrite cell: sheetname='sheet1' rowx=0 colx=0
                    # ��Ҫ���ϣ�cell_overwrite_ok=True)
                    # ������Ϊ�ظ�����һ����Ԫ���µ�

                    sheet.col(0).width = 2 * 256
                    sheet.col(1).width = 19 * 256
                    sheet.col(2).width = 9 * 256
                    sheet.col(3).width = 14 * 256
                    sheet.col(4).width = 19 * 256
                    sheet.col(5).width = 14 * 256
                    for withy_index in range(6, 55):
                        sheet.col(withy_index).width = 12 * 256

                    for row1 in range(2, 12):
                        tall_style = xlwt.easyxf('font:height 330;')  # 36pt,����С�����ֺ�
                        first_row = sheet.row(row1)
                        first_row.set_style(tall_style)

                    for row2 in range(14, 109):
                        tall_style = xlwt.easyxf('font:height 260;')  # 36pt,����С�����ֺ�
                        first_row = sheet.row(row2)
                        first_row.set_style(tall_style)

                    sheet.write(0, 1, shop_name, style11)
                    sheet.write(0, 2, u'', style12)
                    sheet.write(2, 1, u'һ�����̻�����Ϣ', style7)

                    sheet.write_merge(3, 3, 1, 2, u'1��������Ϣ', style8)
                    sheet.write_merge(3, 3, 4, 5, u'2������������Ϣ�����Ϊ��˰ֵ��', style8)
                    sheet.write_merge(3, 3, 7, 10, u'3�����ع�����Ϣ(���Ϊδ˰ֵ��', style8)
                    # sheet.write_merge()
                    stall_info = ['�ٵ�������������', '��ǩԼ������O��', '�۵���������O��', '��װ��������O��', '�ݳ����ʣ�%��', '', '', '']
                    stall_data = [int(row[2]), row[3], row[4], row[5], str(round(float(row[6]), 3) * 100) + '%', '', '',
                                  '']
                    buy_info = ["�ٷ��⣨����ҵ����Ԫ/�£�", "��Ѻ���Ԫ��", "�����������ȣ�%��", "�ܸ��ʽ", "�����޿�ʼ��", "�����޽�ֹ��", "�߿�ҵ����",
                                "�������ڽ�ֹ��"]
                    buy_data = [row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14]]
                    project_info1 = ['�ٻ���װ��', '�ڵ�������', '�����̹���', '����������', '�ݿյ�', '��ȼ��', '', '']
                    project_data1 = [row[16], row[17], row[18], row[19], row[20], row[21], '', '']
                    project_info2 = ['�����簲��', '��ת�÷�', '������', '������', '11.��Ŀ����', '12.���������', '13.��ҵǰ����', '']
                    project_data2 = [row[22], row[23], row[24], row[25], row[26], row[27], row[28], '']

                    shop_basic_resulte = zip(stall_info, stall_data, buy_info, buy_data, project_info1, project_data1,
                                             project_info2, project_data2)
                    i = 4
                    for row3 in shop_basic_resulte:
                        # ��������
                        stall_info = row3[0]
                        stall_data = row3[1]
                        # ���⣨����ҵ����Ԫ/�£�
                        buy_info = row3[2]
                        buy_data = row3[3]
                        # ����װ��
                        project_info1 = row3[4]
                        project_data1 = row3[5]
                        # ���簲��
                        project_info2 = row3[6]
                        project_data2 = row3[7]

                        sheet.write(i, 1, u'%s' % stall_info, style1)
                        sheet.write(i, 2, u'%s' % stall_data, style2)
                        sheet.write(i, 4, u'%s' % buy_info, style1)
                        sheet.write(i, 5, u'%s' % buy_data, style2)
                        sheet.write(i, 7, u'%s' % project_info1, style1)
                        sheet.write(i, 8, u'%s' % project_data1, style2)
                        sheet.write(i, 9, u'%s' % project_info2, style1)
                        sheet.write(i, 10, u'%s' % project_data2, style2)
                        i += 1

                    sheet.write(12, 1, u'', style9)
                    sheet.write(12, 2, u'', style9)
                    sheet.write(12, 4, u'�᷿����ʧ��Ԫ��', style13)
                    sheet.write(12, 5, row[15], style14)

                    sheet.write_merge(11, 12, 7, 9, u'������Ͷ�루Ԫ��', style15)
                    sheet.write_merge(11, 12, 10, 10, row[29], style16)

                    sheet.write(14, 1, u'�ġ��̻����˳���Ϣ', style7)
                    sheet.write(25, 1, u'���������ֽ�����Ϣ', style7)
                    sheet.write(44, 1, u'�������̵��������Ϣ', style7)

                    sheet.write(15, 1, u'����ڼ�', style10)
                    sheet.write(16, 1, u'Ӫҵʱ��', style10)

                    write_index = 2
                    for row4 in results1:
                        # �ֶ�¼�벿��
                        sheet.write(15, write_index, str(row4[1]), style17)
                        sheet.write(16, write_index, '', style17)
                        sheet.write(21, write_index, row4[7], style18)
                        sheet.write(22, write_index, row4[8], style18)
                        sheet.write(23, write_index, row4[9], style18)
                        sheet.write(20, write_index, row4[6], style19)
                        sheet.write(18, write_index, row4[4], style4)
                        sheet.write(19, write_index, row4[5], style4)

                        # �ֶ�¼��-����ɫ��������
                        list22 = [row4[3], row4[10], row4[11], row4[25], row4[26]]
                        list2_index2 = [17, 26, 33, 41, 42]
                        list2_zip2 = zip(list22, list2_index2)
                        for row41 in list2_zip2:
                            sheet.write(row41[1], write_index, u'%s' % row41[0], style20)
                        # �ֶ�¼��-�ޱ�����ɫ
                        list41 = [row4[4], row4[5], row4[11], row4[12], row4[13], row4[14], row4[15], row4[16],
                                  row4[18], row4[19],
                                  row4[20], row4[21], row4[22], row4[23], row4[24]]
                        list4_index1 = [18, 19, 27, 28, 29, 30, 31, 32, 34, 35, 36, 37, 38, 39, 40]
                        list4_zip1 = zip(list41, list4_index1)
                        for row42 in list4_zip1:
                            sheet.write(row42[1], write_index, u'%s' % row42[0],
                                        style4)
                        write_index += 1
                    # ��Դ���ݵõ�����
                    sql_select2 = "SELECT * from `�ŵ������` where `�ŵ�id`='%s'" % row[0]
                    cur2.execute(sql_select2)
                    results2 = cur2.fetchall()

                    sunyibiao_index = 2
                    for row5 in results2:
                        # ��Դ���ݵõ�����-����ɫ�Ӵ�����
                        list51 = [row5[3], row5[10], row5[16], row5[17], row5[18]]
                        list51_index = [45, 52, 58, 59, 60]
                        list51_zip = zip(list51, list51_index)
                        for row51 in list51_zip:
                            sheet.write(row51[1], sunyibiao_index, u'%s' % row51[0], style5)
                        # ��Ӫҵ������
                        sheet.write(46, sunyibiao_index, row5[4], style21)
                        # ���е��
                        sheet.write(50, sunyibiao_index, row5[8], style19)
                        # �������е��
                        sheet.write(56, sunyibiao_index, row5[14], style19)
                        # ��Ӫҵ��ɱ�
                        sheet.write(53, sunyibiao_index, row5[11], style19)
                        # �ۼ�EBITDA
                        sheet.write(77, sunyibiao_index, row5[35], style22)

                        # EBIDTA/��Ͷ��
                        if shop_all_cost2 != 0:
                            formula = row5[33] / shop_all_cost2
                        else:
                            formula = 0
                        sheet.write(76, sunyibiao_index, formula, style6)
                        # ����ɫ��������
                        list52 = [row5[28], row5[29], row5[32], row5[33]]
                        list52_index = [70, 71, 74, 75]
                        list52_zip = zip(list52, list52_index)
                        for row52 in list52_zip:
                            sheet.write(row52[1], sunyibiao_index, u'%s' % row52[0],
                                        style6)
                        # �ޱ�����ɫ
                        list53 = [row5[5], row5[6], row5[7], row5[9],
                                  row5[12], row5[13], row5[15], row5[19], row5[20], row5[21], row5[22], row5[23],
                                  row5[24],
                                  row5[25],
                                  row5[26], row5[27], row5[30], row5[31]]
                        list53_index = [47, 48, 49, 51, 54, 55, 57, 61, 62, 63, 64, 65, 66, 67, 68, 69, 72, 73]
                        list53_zip = zip(list53, list53_index)
                        for row53 in list53_zip:
                            sheet.write(row53[1], sunyibiao_index, u'%s' % row53[0],
                                        style3)

                        sunyibiao_index += 1

                    # ����ɫб������
                    list1 = ['��פ��', '��Ӫҵ��ë��', '����ϱ���']
                    list1_index = 21
                    for row in list1:
                        sheet.write(list1_index, 1, u'%s' % row,
                                    set_style('����', 0, 8, False, True, 43, 0x01, 0x01, 0, 0, 0, 0, True))
                        list1_index += 1
                    # ����ɫ��������
                    list2 = ['����ϵ�����', '�ֽ�����', '�ֽ�����', '���ֽ���', '�ۼƾ��ֽ���', '����Ӫҵ���󣨿����ԡ�-�������У�',
                             '�ۼ�Ӫҵ����', '�����ܶ�����ܶ��ԡ�-�������У�', 'EBIDTA', 'EBIDTA/��Ͷ��']
                    list2_index = [17, 26, 33, 41, 42, 70, 71, 74, 75, 76]
                    list2_zip = zip(list2, list2_index)
                    for row in list2_zip:
                        sheet.write(row[1], 1, u'%s' % row[0], style6)
                    # ����ɫ�Ӵ�����
                    list3 = ['һ��Ӫҵ����', '����Ӫҵ�ɱ�', 'ë����', 'ë���ʣ�%��', '���۷���']
                    list3_index = [45, 52, 58, 59, 60]
                    list3_zip = zip(list3, list3_index)
                    for row in list3_zip:
                        sheet.write(row[1], 1, u'%s' % row[0], style5)

                    sheet.write(46, 1, u'��Ӫҵ������', style21)
                    sheet.write(20, 1, u'���У�Ѻ��ת���루����', style19)
                    sheet.write(50, 1, u'���е��', style19)
                    sheet.write(56, 1, u'���е��', style19)
                    sheet.write(53, 1, u'��Ӫҵ��ɱ�', style19)

                    sheet.write(77, 1, u'�ۼ�EBITDA', style22)

                    # �ޱ�����ɫ
                    list4 = ["����������", "�˳�������", "�ٷ�������", "��Ѻ������", "��ʳƷ��ȫ��֤������", "�ܽ���������", "����Դ������", "����������", "�ٳ�����ҵ����",
                             "��װ������",
                             "����Դ������", "���̻��˿�����", "����Ա�ɱ�����", "��ά������", "����������", "���̻��������", "���̻�����������", "����Դ������", "����������",
                             "�ٳ�����ҵ��",
                             "����ҵ�ѳɱ�", "��������سɱ�", "��Ա��н��", "�ڰ칫�ѽ�ͨ�д���", "���۾ɷ�", "�ܿ����̯��", "��ά�޷�", "����ҵ���շ�", "����ѯ�������",
                             "���ֵ�׺�Ʒ",
                             "����������", "Ӫҵ������", "Ӫҵ��֧��"]
                    list4_index = [18, 19, 27, 28, 29, 30, 31, 32, 34, 35, 36, 37, 38, 39, 40, 47, 48, 49, 51, 54, 55,
                                   57, 61, 62,
                                   63, 64, 65, 66, 67, 68, 69, 72, 73]
                    list4_zip = zip(list4, list4_index)
                    for row in list4_zip:
                        sheet.write(row[1], 1, u'%s' % row[0], style3)
                path='C:\\Users\\tl\\PycharmProjects\\get_excel\\finance\\shop_card-%s.xlsx' % int(time.time())
                book.save(path)


                filename = os.path.basename(path)
                response = Response(file_iterator(path))
                response.headers['Content-Type'] = 'application/octet-stream'
                response.headers["Content-Disposition"] = 'attachment;filename="{}"'.format(filename)
                return response
#�ϴ��ŵ꿨Ƭ��ͼ
@app.route('/text',methods=['POST'])
def text():

    file = request.files.get('fafafa')
    if file is None:
        return "δ�ϴ��ļ���"
    wb = load_workbook(file, data_only=True)
    sheetnames = wb.sheetnames
    for sheet_name in sheetnames:

        ws = wb[sheet_name]
        shop_name2 = sheet_name

        shop_id = str(uuid.uuid3(uuid.NAMESPACE_DNS, str(shop_name2)))

        # �ŵ���Ϣ
        baic_list = []
        baic_list.extend([shop_id, shop_name2])
        baic_col_c = [5, 6, 7, 8, 9]
        for col_c in baic_col_c:
            c = ws['C%s' % col_c].value
            if c is None or c == '#VALUE!':
                c = 0

            if (str(type(c)) =="<class 'float'>" ):
                if c>1.0:
                    c=round(c,0)
            baic_list.append(c)

        baic_col_f = [5, 6, 7, 8, 9, 10, 11, 12, 13]
        for col_f in baic_col_f:
            c1 = ws['F%s' % col_f].value
            if c1 is None:
                c1 = 0
            if (str(type(c1)) == "<class 'float'>"):
                if c1 > 1.0:
                    c1 = round(c1, 0)
            baic_list.append(c1)

        baic_col_i = [5, 6, 7, 8, 9, 10]
        for col_i in baic_col_i:
            c2 = ws['I%s' % col_i].value
            if c2 is None:
                c2 = 0
            if (str(type(c2)) != "<class 'str'>"):
                if c2 > 1.0:
                    c2 = round(c2, 0)
            baic_list.append(c2)

        baic_col_k = [5, 6, 7, 8, 9, 10, 11, 12]
        for col_k in baic_col_k:
            c3 = ws['K%s' % col_k].value
            if c3 is None:
                c3 = 0
            if (str(type(c3)) == "<class 'float'>"):
                if c3 > 1.0:
                    c3 = round(c3, 0)
            baic_list.append(c3)

        row = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
               'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
               'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY']
        luru_list = []

        sunyi_list = []

        luru_col = [16, 18, 19, 20, 21, 22, 23, 24, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43]
        sunyi_col = [16, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69,
                     70, 71, 72, 73, 74, 75, 76, 77, 78]

        test_col=[22, 23,24]
        for row_test in row:
            for test_col1 in test_col:
                data_luru2  = ws[row_test + str(test_col1)].value




        for row2 in row:
            luru_list2 = []
            sunyi_list2 = []
            luru_list2.extend([shop_id, shop_name2])
            sunyi_list2.extend([shop_id, shop_name2])
            for luru_col2 in luru_col:
                # ¼����Ϣ
                data_luru  = ws[row2 + str(luru_col2)].value
                if data_luru is None:
                    data_luru = 0

                if (str(type(data_luru)) !="<class 'datetime.datetime'>"):
                    if (luru_col2==24):
                        data_luru = round(data_luru, 2)

                    elif (data_luru >= 1.0):
                        data_luru = round(data_luru, 0)
                    elif (data_luru < 1.0):
                        data_luru = round(data_luru, 2)


                luru_list2.append(data_luru)
            for sunyi_col2 in sunyi_col:
                data_sunyi = ws[row2 + str(sunyi_col2)].value
                if data_sunyi is None:
                    data_sunyi = 0
                if (str(type(data_sunyi)) == "<class 'float'>" or str(type(data_sunyi)) == "<class 'int'>"):
                    if (sunyi_col2==77):
                        data_sunyi = round(data_sunyi, 3)
                    elif (data_sunyi >= 1.0):
                        data_sunyi = round(data_sunyi, 0)
                    elif (data_sunyi < 1.0):
                        data_sunyi = round(data_sunyi, 2)
                sunyi_list2.append(data_sunyi)

            luru_list.append(tuple(luru_list2))
            sunyi_list.append(tuple(sunyi_list2))

        sql_select = "INSERT ignore into `�ŵ������`  (`�ŵ�id`,`�ŵ�����`,`����ڼ�`,`Ӫҵ����`,`��Ӫҵ������`,`�̻��������`,`�̻�����������`,`��Դ������`,`���е��`,`��������`,`����Ӫҵ�ɱ�`,`��Ӫҵ��ɱ�`,`������ҵ��`,`��ҵ�ѳɱ�`,`�������`,`������سɱ�`,`ë����`,`ë����`,`���۷���`,`Ա��н��`,`�칫�ѽ�ͨ�д���`,`�۾ɷ�`,`�����̯��`,`ά�޷�`,`��ҵ���շ�`,`��ѯ�������`,`��ֵ�׺�Ʒ`,`��������`,`Ӫҵ���󣨿����ԡ�-�������У�`,`�ۼ�Ӫҵ����`,`Ӫҵ������`,`Ӫҵ��֧��`,`�����ܶ�����ܶ��ԡ�-�������У�`,`EBIDTA`,`EBIDTA/��Ͷ��`,`�ۼ�EBITDA`) VALUES "
        sql_select2 = "INSERT ignore into `�ŵ���Ϣ`  (`�ŵ�id`,`�ŵ�����`,`��������������`,`ǩԼ�����ƽ���ף�`,`���������ƽ���ף�`,`װ��������O��`,`������`,`���⣨����ҵ����Ԫ/�£�`,`Ѻ���Ԫ��`,`���������ȣ�%��`,`���ʽ`,`���޿�ʼ��`,`���޽�����`,`��ҵʱ��`,`�����ֹ��`,`������ʧ��Ԫ��`,`����װ��`,`��������`,`���̹���`,`��������`,`�յ�`,`ȼ��`,`���簲��`,`ת�÷�`,`����`,`����`,`��Ŀ����`,`���������`,`��ҵǰ����`,`������Ͷ�루Ԫ��`)   VALUES "
        sql_select3 = "INSERT ignore into `��������¼���` (`�ŵ�id`,`�ŵ�����`,`����ڼ�`,`����ϵ�����`,`����`,`�˳�`,`Ѻ��ת����`,`��ס��`,`��Ӫҵ��ë��`,`����ϱ���`,`�ֽ�����`,`��������`,`Ѻ������`,`ʳƷ��ȫ��֤������`,`����������`,`��Դ������`,`��������`,`�ֽ�����`,`������ҵ����`,`װ������`,`��Դ������`,`�̻��˿�����`,`��Ա�ɱ�����`,`ά������`,`��������`,`���ֽ���`,`�ۼƾ��ֽ���`)    VALUES "

        db = pymysql.connect(host="bj-cdb-cwu7v42u.sql.tencentcdb.com", user="root",
                             password="xmxc1234", db="test", port=62864)
        cur = db.cursor()

        for data in sunyi_list:
            sql_select += " ('%s','%s','%s',%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%s,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%s,%d)," % data

        sql_select2 += " ('%s','%s',%d,'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d)," % tuple(
            baic_list)

        for data3 in luru_list:
            sql_select3 += " ('%s','%s','%s',%d,%d,%d,%d,%s,%s,%s,%d,%d,%d,%d,%d,%d,%d,%s,%d,%d,%d,%d,%d,%d,%d,%d,%d)," % data3

        sql_select = sql_select.rstrip(',')
        sql_select2 = sql_select2.rstrip(',')
        sql_select3 = sql_select3.rstrip(',')

        cur.execute(sql_select)
        cur.execute(sql_select2)
        cur.execute(sql_select3)
        db.commit()  # �����ύ
        cur.close()
        db.close()


        # try:
        #     cur.execute(sql_select)
        #     cur.execute(sql_select2)
        #     cur.execute(sql_select3)
        # except Exception as e:
        #     db.rollback()  # ����ع�
        # else:
        #     db.commit()  # �����ύ
        #     cur.close()
        #     db.close()
    return '�ϴ��ɹ���'
#�ϴ�Դ������ͼ
@app.route('/upload_source_data',methods=['POST'])
def upload_source_data():

    file = request.files.get('source_data_id')
    if file is None:
        return "δ�ϴ��ļ���"
    wb = load_workbook(file)

    sheetnames = wb.sheetnames

    # �ۼ�Ӫҵ����
    all_data = []
    for sheet_name in sheetnames:

        add_opening_profit = 0.0
        add_ebi = 0.0
        ws = wb[sheet_name]
        shop_id = uuid.uuid3(uuid.NAMESPACE_DNS, str(sheet_name))

        # cell_range = ws['A1':'C2']
        max_column = ws.max_column
        for column_index in range(3, 15):
            shop_list = []
            shop_list.append(str(shop_id))

            # ����ڼ�
            opening_time = ws.cell(row=2, column=column_index).value
            # opening_time='2019-'
            shop_list.append(opening_time)
            shop_list.append(sheet_name)
            # Ӫҵ����
            opening_income = float(ws.cell(row=3, column=column_index).value)
            shop_list.append(opening_income)
            # ��Ӫҵ������
            main_income = ws.cell(row=4, column=column_index).value
            shop_list.append(main_income)
            # �̻��������
            shop_rent_income = ws.cell(row=5, column=column_index).value
            shop_list.append(shop_rent_income)
            # �̻�����������
            shop_come_income = ws.cell(row=6, column=column_index).value
            shop_list.append(shop_come_income)
            # ��Դ������
            energy_income = 0.0
            for i in range(8, 14):
                energy_income += float(ws.cell(row=i, column=column_index).value)
            shop_list.append(energy_income)
            # ���е������
            in_electric = ws.cell(row=9, column=column_index).value
            shop_list.append(in_electric)
            # ��������
            other_income = 0.0
            for i in range(17, 24):
                if ws.cell(row=i, column=column_index).value is None:
                    tem = 0
                else:
                    tem = ws.cell(row=i, column=column_index).value
                other_income += float(tem)
            shop_list.append(other_income)
            # Ӫҵ�ɱ�
            subtract_cost = ws.cell(row=24, column=column_index).value
            shop_list.append(subtract_cost)
            # ������ҵ��
            changzu_cost = float(ws.cell(row=25, column=column_index).value)

            # ��ҵ�ѳɱ�
            wuye_cost = float(ws.cell(row=26, column=column_index).value)

            # ��Ӫҵ��ɱ�
            main_cost = changzu_cost + wuye_cost
            shop_list.append(main_cost)
            shop_list.append(changzu_cost)
            shop_list.append(wuye_cost)
            # ���е�ѳɱ�
            energy_cost = ws.cell(row=29, column=column_index).value
            shop_list.append(energy_cost)
            # �����ɱ�
            other_cost = ws.cell(row=27, column=column_index).value
            shop_list.append(other_cost)
            # ë����
            maolie = opening_income - subtract_cost
            shop_list.append(maolie)
            # ë����
            if opening_income != 0:
                maoli_lu = maolie / opening_income
                maoli_lu = str(maoli_lu)
            else:
                maoli_lu = 0
                maoli_lu = str(maoli_lu)

            shop_list.append(maoli_lu)
            # ���۷���
            shopping_money = ws.cell(row=38, column=column_index).value
            shop_list.append(shopping_money)
            # Ա��н��
            people_money = ws.cell(row=65, column=column_index).value
            shop_list.append(people_money)
            # �칫�ѽ�ͨ�д���
            office = 0.0
            for i in range(43, 48):
                office += float(ws.cell(row=i, column=column_index).value)
            office += float(ws.cell(row=50, column=column_index).value)
            shop_list.append(office)
            # �۾ɷ�
            old_cost = ws.cell(row=60, column=column_index).value
            shop_list.append(old_cost)
            # �����̯��
            kaiban_cost = ws.cell(row=51, column=column_index).value
            shop_list.append(kaiban_cost)
            # ά�޷�
            maintain_cost = ws.cell(row=49, column=column_index).value
            shop_list.append(maintain_cost)
            # ��ҵ���շ�
            insurance_cost = ws.cell(row=52, column=column_index).value
            shop_list.append(insurance_cost)
            # ��ѯ�������
            service_cost = ws.cell(row=57, column=column_index).value
            shop_list.append(service_cost)
            # ��ֵ�׺�Ʒ
            yihaopin_cost = ws.cell(row=54, column=column_index).value
            shop_list.append(yihaopin_cost)
            # ��������
            other_money = ws.cell(row=53, column=column_index).value + ws.cell(row=55,
                                                                               column=column_index).value + ws.cell(
                row=56, column=column_index).value + ws.cell(row=58, column=column_index).value + ws.cell(row=59,
                                                                                                          column=column_index).value
            shop_list.append(other_money)
            # ����Ӫҵ���󣨿����ԡ�-�������У�
            opening_profit = float(ws.cell(row=104, column=column_index).value)
            shop_list.append(opening_profit)
            add_opening_profit += opening_profit
            shop_list.append(add_opening_profit)
            # Ӫҵ������
            out_shopping_income = ws.cell(row=105, column=column_index).value
            shop_list.append(out_shopping_income)
            # Ӫҵ��֧��
            out_shopping_cost = ws.cell(row=106, column=column_index).value
            shop_list.append(out_shopping_cost)
            # �����ܶ�����ܶ��ԡ�-�������У�
            all_profit = ws.cell(row=107, column=column_index).value
            shop_list.append(all_profit)
            # EBIDTA
            ebidta = opening_profit + old_cost + kaiban_cost
            ebidta2 = "'%d/$K$12'" % ebidta
            shop_list.append(ebidta)
            shop_list.append(ebidta2)
            add_ebi += ebidta
            shop_list.append(add_ebi)
            all_data.append(tuple(shop_list))

    sql_select = "INSERT IGNORE into `�ŵ������`  (`�ŵ�id`,`����ڼ�`,`�ŵ�����`,`Ӫҵ����`,`��Ӫҵ������`,`�̻��������`,`�̻�����������`,`��Դ������`,`���е��`,`��������`,`����Ӫҵ�ɱ�`,`��Ӫҵ��ɱ�`,`������ҵ��`,`��ҵ�ѳɱ�`,`�������`,`������سɱ�`,`ë����`,`ë����`,`���۷���`,`Ա��н��`,`�칫�ѽ�ͨ�д���`,`�۾ɷ�`,`�����̯��`,`ά�޷�`,`��ҵ���շ�`,`��ѯ�������`,`��ֵ�׺�Ʒ`,`��������`,`Ӫҵ���󣨿����ԡ�-�������У�`,`�ۼ�Ӫҵ����`,`Ӫҵ������`,`Ӫҵ��֧��`,`�����ܶ�����ܶ��ԡ�-�������У�`,`EBIDTA`,`EBIDTA/��Ͷ��`,`�ۼ�EBITDA`) VALUES "
    sql_select2 = "INSERT IGNORE into `�ŵ���Ϣ`  (`�ŵ�id`,`�ŵ�����`) VALUES "
    sql_select3 = "INSERT IGNORE into `��������¼���`  (`�ŵ�id`,`����ڼ�`,`�ŵ�����`) VALUES "

    db = pymysql.connect(host="bj-cdb-cwu7v42u.sql.tencentcdb.com", user="root",
                         password="xmxc1234", db="test", port=62864)
    cur = db.cursor()

    for data in all_data:
        id = data[0]
        shop_name = data[2]

        shop_date = data[1]

        sql_select += " ('%s','%s','%s',%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%s,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%d,%s,%d)," % data
        sql_select2 += " ('%s','%s')," % (id, shop_name)
        sql_select3 += " ('%s','%s','%s')," % (id, shop_date, shop_name)

    sql_select = sql_select.rstrip(',')
    sql_select2 = sql_select2.rstrip(',')
    sql_select3 = sql_select3.rstrip(',')

    try:
        cur.execute(sql_select)
        cur.execute(sql_select2)
        cur.execute(sql_select3)
    except Exception as e:
        db.rollback()  # ����ع�
    else:
        db.commit()  # �����ύ
        cur.close()
        db.close()

    return '�ϴ��ɹ���'

if __name__ == '__main__':
    app.run(debug=True)
